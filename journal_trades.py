"""Add trades from a CSV file into the 'OPTIONS DEMO.xlsx' journal.

The script looks for a CSV file named ``recent trades.csv``. If it does not
exist it falls back to ``all_trades.csv``. Each trade is added as a new row in
the workbook using the same formulas and formatting as the existing rows.

Run the script with ``python journal_trades.py``.
"""
from __future__ import annotations

import csv
from copy import copy
from datetime import datetime
import math
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook


def _norm_pdf(x: float) -> float:
    """Probability density function of the standard normal distribution."""
    return math.exp(-0.5 * x * x) / math.sqrt(2 * math.pi)


def _norm_cdf(x: float) -> float:
    """Cumulative distribution function for the standard normal distribution."""
    return 0.5 * (1 + math.erf(x / math.sqrt(2)))


def _greeks(option: str, s: float, k: float, t: float, sigma: float, qty: float) -> tuple[Optional[float], Optional[float], Optional[float], Optional[float]]:
    """Return delta, gamma, theta and vega for an option position."""
    if None in (s, k, t, sigma, qty) or t <= 0 or sigma <= 0:
        return None, None, None, None
    d1 = (math.log(s / k) + 0.5 * sigma ** 2 * t) / (sigma * math.sqrt(t))
    d2 = d1 - sigma * math.sqrt(t)
    pdf = _norm_pdf(d1)
    if option == "CALL":
        delta = _norm_cdf(d1)
        theta = -(s * pdf * sigma) / (2 * math.sqrt(t))
    else:
        delta = _norm_cdf(d1) - 1
        theta = -(s * pdf * sigma) / (2 * math.sqrt(t))
    gamma = pdf / (s * sigma * math.sqrt(t))
    vega = s * pdf * math.sqrt(t) / 100
    # Convert theta to per-day and scale all Greeks by position size.
    theta /= 365
    return delta * qty, gamma * qty, theta * qty, vega * qty


def _csv_path() -> Path:
    """Return the path to the CSV file containing the trade data."""
    for name in ("recent trades.csv", "all_trades.csv"):
        path = Path(name)
        if path.exists():
            return path
    raise FileNotFoundError("No trade CSV file found.")


def _float(value: str) -> Optional[float]:
    """Convert ``value`` to float if possible."""
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _parse_trade_logs() -> dict[str, dict[str, float]]:
    """Return a mapping of symbol to data parsed from trade log text files.

    The trading script generates log files named ``option_trade_log_*.log``
    (or ``*.txt``) that contain a ``Ticker Data`` section with market data and a
    ``Greek Exposures`` section with the per-position Greeks.  This function
    scans the current directory for such files and builds a dictionary keyed by
    symbol with any numeric values found.  If no log files are present an empty
    dictionary is returned.
    """

    logs: dict[str, dict[str, float]] = {}
    for path in Path().glob("option_trade_log_*.*"):
        symbol: Optional[str] = None
        ticker: dict[str, float] = {}
        greeks: dict[str, float] = {}
        section: Optional[str] = None
        for raw in path.read_text().splitlines():
            line = raw.strip()
            if not line:
                continue
            if line.startswith("Placing"):
                parts = line.split()
                if len(parts) >= 4:
                    symbol = parts[3]
            elif line.startswith("symbol:"):
                symbol = line.split(":", 1)[1].strip()
            elif line == "Ticker Data:":
                section = "ticker"
            elif line == "Greek Exposures:":
                section = "greeks"
            elif section == "ticker" and ":" in line:
                key, val = line.split(":", 1)
                value = _float(val.strip())
                if value is not None:
                    ticker[key.strip()] = value
            elif section == "greeks" and not line.lower().startswith("greek"):
                parts = line.split()
                if len(parts) >= 2:
                    greek = parts[0].lower()
                    exposure = _float(parts[-1].rstrip("."))
                    if exposure is not None:
                        greeks[greek] = exposure
        if symbol:
            logs[symbol] = {**ticker, **greeks}
    return logs

def _parse_symbol(symbol: str) -> tuple[str, datetime, float, str]:
    """Return (pair, expiry, strike, option_type) extracted from ``symbol``."""
    base, expiry_str, strike, opt, *_ = symbol.split("-")
    expiry = datetime.strptime(expiry_str, "%d%b%y")
    option_type = "CALL" if opt.upper().startswith("C") else "PUT"
    return base, expiry, float(strike), option_type


def _money(option: str, strike: float, underlying: Optional[float]) -> Optional[str]:
    """Return 'ITM' or 'OTM' if ``underlying`` price is available."""
    if underlying is None:
        return None
    if option == "CALL":
        return "ITM" if underlying >= strike else "OTM"
    return "ITM" if underlying <= strike else "OTM"


def main() -> None:
    csv_file = _csv_path()
    wb = load_workbook("OPTIONS DEMO.xlsx")
    ws = wb.active

    # Use row 3 as a template for formatting.
    template_row = 3
    max_col = ws.max_column

    # Find the last row with data in column B (Option column).
    last_row = ws.max_row
    while last_row > 1 and ws.cell(row=last_row, column=2).value is None:
        last_row -= 1

    with csv_file.open(newline="") as f:
        trades = list(csv.DictReader(f))

    # Parse any trade log files for additional data such as Greeks.
    log_data = _parse_trade_logs()

    # Sort trades by time so entries appear before exits.
    trades.sort(key=lambda t: datetime.strptime(t["localTime"], "%d/%m/%Y %H:%M"))

    open_trades: dict[str, dict] = {}
    for trade in trades:
        symbol = trade["symbol"]
        if symbol in open_trades:
            entry = open_trades.pop(symbol)
            exit_trade = trade

            last_row += 1
            pair, expiry, strike, option_type = _parse_symbol(symbol)
            entry_time = datetime.strptime(entry["localTime"], "%d/%m/%Y %H:%M")
            exit_time = datetime.strptime(exit_trade["localTime"], "%d/%m/%Y %H:%M")
            qty = _float(entry.get("execQty"))
            log = log_data.get(symbol, {})
            underlying = _float(entry.get("underlyingPrice")) or log.get("underlyingPrice")
            iv = (
                _float(entry.get("tradeIv") or entry.get("markIv") or exit_trade.get("tradeIv") or exit_trade.get("markIv"))
                or log.get("markIv")
            )
            t_expiry = (expiry - entry_time).total_seconds() / (365 * 24 * 60 * 60)
            delta = log.get("delta")
            gamma = log.get("gamma")
            theta = log.get("theta")
            vega = log.get("vega")
            if any(g is None for g in (delta, gamma, theta, vega)):
                delta, gamma, theta, vega = _greeks(option_type, underlying, strike, t_expiry, iv, qty or 0)

            # Fees and result
            fee = (_float(entry.get("netFee") or entry.get("execFee")) or 0) + (
                _float(exit_trade.get("netFee") or exit_trade.get("execFee")) or 0
            )
            result = "WIN" if _float(exit_trade.get("balance")) > _float(entry.get("balance")) else "LOSS"

            # Basic values
            ws.cell(row=last_row, column=1, value=_money(option_type, strike, underlying))
            ws.cell(row=last_row, column=2, value=option_type)
            ws.cell(row=last_row, column=3, value=strike)
            ws.cell(row=last_row, column=4, value=pair)
            ws.cell(row=last_row, column=5, value=(entry.get("orderType") or "").upper())
            ws.cell(row=last_row, column=6, value=(entry.get("side") or "").upper())
            ws.cell(row=last_row, column=8, value=_float(entry.get("execPrice")))
            ws.cell(row=last_row, column=9, value=_float(exit_trade.get("execPrice")))
            ws.cell(row=last_row, column=10, value=qty)
            ws.cell(row=last_row, column=11, value=delta)
            ws.cell(row=last_row, column=12, value=gamma)
            ws.cell(row=last_row, column=13, value=theta)
            ws.cell(row=last_row, column=14, value=vega)
            ws.cell(row=last_row, column=15, value=entry_time)
            ws.cell(row=last_row, column=16, value=exit_time)
            ws.cell(row=last_row, column=18, value=expiry)
            ws.cell(row=last_row, column=20, value=fee)
            ws.cell(row=last_row, column=22, value=iv)
            entry_index = _float(entry.get("indexPrice")) or log.get("indexPrice")
            exit_index = _float(exit_trade.get("indexPrice")) or log.get("indexPrice")
            ws.cell(row=last_row, column=23, value=entry_index)
            ws.cell(row=last_row, column=24, value=exit_index)
            ws.cell(row=last_row, column=25, value=None)
            ws.cell(row=last_row, column=30, value=result)

            # Formulas
            ws.cell(row=last_row, column=7, value=f'=IF(F{last_row}="BUY", (H{last_row}*J{last_row})/-1,(H{last_row}*J{last_row}))')
            ws.cell(row=last_row, column=17, value=f'=TEXT((P{last_row}-O{last_row}), "hh:mm")')
            ws.cell(row=last_row, column=19, value=f'=R{last_row}-O{last_row}')
            ws.cell(row=last_row, column=27, value=f'=U{last_row}+T{last_row}+G{last_row}+(I{last_row}/100)')
            ws.cell(row=last_row, column=28, value=f'=AA{last_row}/AC{last_row-1}')
            ws.cell(row=last_row, column=29, value=f'=AC{last_row-1}+AA{last_row}')

            # Copy formatting from the template row.
            for col in range(1, max_col + 1):
                tmpl = ws.cell(row=template_row, column=col)
                cell = ws.cell(row=last_row, column=col)
                cell.font = copy(tmpl.font)
                cell.fill = copy(tmpl.fill)
                cell.border = copy(tmpl.border)
                cell.alignment = copy(tmpl.alignment)
                cell.number_format = tmpl.number_format
        # Ignore any unmatched trades.

    wb.save("OPTIONS DEMO.xlsx")


if __name__ == "__main__":
    main()
